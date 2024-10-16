import time
import traceback

from sedona.spark import *
import pyspark.sql.functions as f
from pyspark.sql import SparkSession

# Get active SparkSession
spark = SparkSession.getActiveSession()

# Test functions
def test_numpy():
    import numpy as np

    a = np.array([1, 2, 3])
    assert a.sum() == 6
    print("NumPy test passed.")


def test_openpyxl():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["B1"] = "World"
    assert ws["A1"].value == "Hello"
    print("openpyxl test passed.")


def test_pint():
    from pint import UnitRegistry

    ureg = UnitRegistry()
    distance = 24.0 * ureg.meter
    time_duration = 8.0 * ureg.second
    speed = distance / time_duration
    assert speed.magnitude == 3.0
    assert str(speed.units) == "meter / second"
    print("Pint test passed.")


def test_typing_extensions():
    from typing_extensions import TypedDict

    class Point2D(TypedDict):
        x: int
        y: int

    p: Point2D = {"x": 1, "y": 2}
    assert p["x"] == 1
    print("typing_extensions test passed.")


def test_colorama():
    from colorama import init, Fore

    init()
    print(Fore.RED + "Colorama test passed.")


def test_shapely():
    from shapely.geometry import Point

    p1 = Point(0, 0)
    p2 = Point(1, 1)
    line = p1.buffer(1).intersection(p2.buffer(1))
    assert line.area > 0
    print("Shapely test passed.")


def test_pyogrio():
    import pyogrio

    drivers = pyogrio.list_drivers()
    assert "ESRI Shapefile" in drivers
    print("pyogrio test passed.")


def test_geopandas():
    import geopandas as gpd
    from shapely.geometry import Point

    data = {"geometry": [Point(1, 1), Point(2, 2)]}
    gdf = gpd.GeoDataFrame(data, crs="EPSG:4326")
    assert len(gdf) == 2
    print("GeoPandas test passed.")


def test_rasterio():
    import rasterio

    print("rasterio version:", rasterio.__version__)
    print("rasterio test passed.")


def test_midgard():
    import midgard
    from midgard.data.position import Position
    from midgard.math.unit import Unit
    from midgard.math import ellipsoid

    print("Midgard version:", midgard.__version__)
    print("Midgard test passed.")


def test_dask_felleskomponenter():
    import dask_felleskomponenter

    print("dask-felleskomponenter imported successfully.")
    print("dask_felleskomponenter test passed.")


def test_pynex():
    from pynex.pynex import Pynex

    print("pynex test passed.")


def test_spark():
    df = spark.createDataFrame([(1, "a"), (2, "b")], ["id", "value"])
    assert df.count() == 2

    print("Apache Spark test passed.")


def test_pyspark_sedona():
    # Create a coordinate (longitude, latitude)
    point = [(10.75, 59.9167)]  # Oslo, Norway
    # Create a DataFrame with coordinate columns containing the point
    df = spark.createDataFrame(point, ["longitude", "latitude"])
    # Add geometry column
    df = df.withColumn("geometry", f.expr("st_point(longitude, latitude)"))
    # Assert geometry column is of type geometry
    assert df.schema["geometry"].dataType.typeName() == "geometrytype"

    print("Apache Sedona test passed.")


def test_sedona_geopandas():
    import geopandas as gpd
    from shapely.geometry import Point

    data = {"id": [1, 2], "geometry": [Point(0, 0), Point(1, 1)]}
    gdf = gpd.GeoDataFrame(data, crs="EPSG:4326", geometry="geometry")

    sdf = spark.createDataFrame(gdf)
    sdf.createOrReplaceTempView("geom_table")
    result = spark.sql("SELECT ST_Buffer(geometry, 1) as buffer_geom FROM geom_table")
    assert result.count() == 2

    print(
        "Sedona and GeoPandas integration test passed. \
        \nGeometry type produces Arrow warnings, but works with non-optimized fallback."
    )


def test_spark_numpy():
    from pyspark.sql.functions import udf
    from pyspark.sql.types import DoubleType
    import numpy as np

    df = spark.createDataFrame([(1,), (4,), (9,)], ["value"])

    def numpy_sqrt(x):
        return float(np.sqrt(x))

    sqrt_udf = udf(numpy_sqrt, DoubleType())
    df = df.withColumn("sqrt", sqrt_udf(df["value"]))
    result = df.collect()
    assert result[0]["sqrt"] == 1.0

    print("Spark and NumPy integration test passed.")


def test_sedona_kepler_pydeck():
    # Create a Point geometry (longitude, latitude)
    point = [(10.75, 59.9167)]  # Oslo, Norway
    # Create a DataFrame from the point
    df = spark.createDataFrame(point, ["longitude", "latitude"])
    # Add geometry column
    df = df.withColumn("geometry", f.expr("st_point(longitude, latitude)"))

    # Plot in SedonaKepler
    map = SedonaKepler.create_map()
    SedonaKepler.add_df(map, df, name="Oslo Point")
    map

    # Plot in SedonaPyDeck
    SedonaPyDeck.create_geometry_map(
        df,
        fill_color="[85, 183, 177, 255]",
        line_color="[85, 183, 177, 255]",
        elevation_col=0,
        initial_view_state=None,
        map_style=None,
        map_provider=None,
    )


def get_all_tests():
    # Return complete list of test functions
    return [
        test_numpy,
        test_openpyxl,
        test_pint,
        test_typing_extensions,
        test_colorama,
        test_shapely,
        test_pyogrio,
        test_geopandas,
        test_rasterio,
        test_midgard,
        test_dask_felleskomponenter,
        test_pynex,
        test_spark,
        test_pyspark_sedona,
        test_sedona_geopandas,
        test_spark_numpy,
        test_sedona_kepler_pydeck,
    ]


# Function to run a single test function
def run_test(test_func):
    """Runs a single test function, handles timing, exception catching, and printing."""
    test_name = test_func.__name__
    print(f"{'='*60}\nRunning {test_name}...\n{'-'*60}")
    start_time = time.time()
    try:
        test_func()
        elapsed_time = time.time() - start_time
        print(f"\n{test_name} PASSED in {elapsed_time:.2f} seconds.\n")
        return True, test_name, elapsed_time, None
    except Exception as e:
        elapsed_time = time.time() - start_time
        print(f"\n{test_name} FAILED in {elapsed_time:.2f} seconds.\n")
        print("Error details:")
        traceback.print_exc()
        print("\n")
        return False, test_name, elapsed_time, e


# Function to run all tests
def run_all_tests():
    """Runs all test functions and prints summary."""
    test_functions = get_all_tests()
    passed_tests = []
    failed_tests = []
    print("\n" + "=" * 60)
    print("Starting Test Suite")
    print("=" * 60 + "\n")
    total_start_time = time.time()
    for test_func in test_functions:
        success, test_name, elapsed_time, error = run_test(test_func)
        if success:
            passed_tests.append((test_name, elapsed_time))
        else:
            failed_tests.append((test_name, elapsed_time, error))
    total_elapsed_time = time.time() - total_start_time
    # Print summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    print(f"Total tests run: {len(test_functions)}")
    print(f"Tests passed: {len(passed_tests)}")
    print(f"Tests failed: {len(failed_tests)}")
    print(f"Total time: {total_elapsed_time:.2f} seconds")
    if failed_tests:
        print("\nFailed Tests:")
        for test_name, elapsed_time, error in failed_tests:
            print(f"- {test_name}: {error}")
    else:
        print("\nAll tests passed successfully.")


if __name__ == "__main__":
    run_all_tests()
